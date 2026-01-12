"""
Reporting & Export Service
Generates PDF and Excel reports with MinIO storage
"""
from fastapi import FastAPI, Depends, HTTPException, Query, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional, Dict, Any
from uuid import UUID
from datetime import datetime, timedelta
import os
import sys
import json
import io
from decimal import Decimal

sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..'))

from shared.database import create_db_engine, get_session_factory, get_db_session, Base
from shared.auth import get_current_user, require_roles, Roles
from shared.logging_config import configure_logging
from shared.events import EventPublisher

import models
from pydantic import BaseModel, ConfigDict, Field

# Excel and PDF libraries
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from weasyprint import HTML, CSS
import httpx

# MinIO client
from minio import Minio
from minio.error import S3Error

# Initialize
logger = configure_logging("reporting-service")
DATABASE_URL = os.getenv("DATABASE_URL")
engine = create_db_engine(DATABASE_URL)
SessionFactory = get_session_factory(engine)
if os.getenv("AUTO_CREATE_DB", "true").lower() == "true":
    Base.metadata.create_all(bind=engine)

RABBITMQ_URL = os.getenv("RABBITMQ_URL")
event_publisher = EventPublisher(RABBITMQ_URL, "reporting-service") if RABBITMQ_URL else None

# MinIO configuration
MINIO_ENDPOINT = os.getenv("MINIO_ENDPOINT", "minio:9000")
MINIO_ACCESS_KEY = os.getenv("MINIO_ACCESS_KEY") or os.getenv("MINIO_ROOT_USER")
MINIO_SECRET_KEY = os.getenv("MINIO_SECRET_KEY") or os.getenv("MINIO_ROOT_PASSWORD")
MINIO_BUCKET = os.getenv("MINIO_BUCKET", "reports")
MINIO_SECURE = os.getenv("MINIO_SECURE", "false").lower() == "true"

# Initialize MinIO client
if MINIO_ACCESS_KEY and MINIO_SECRET_KEY:
    try:
        minio_client = Minio(
            MINIO_ENDPOINT,
            access_key=MINIO_ACCESS_KEY,
            secret_key=MINIO_SECRET_KEY,
            secure=MINIO_SECURE
        )
        # Create bucket if not exists
        if not minio_client.bucket_exists(MINIO_BUCKET):
            minio_client.make_bucket(MINIO_BUCKET)
            logger.info(f"Created MinIO bucket: {MINIO_BUCKET}")
    except Exception as e:
        logger.warning(f"MinIO initialization failed: {e}. Reports will not be stored.")
        minio_client = None
else:
    logger.warning("MinIO credentials not set; reports will not be stored.")
    minio_client = None

# Service URLs for data fetching
INVENTORY_SERVICE_URL = os.getenv("INVENTORY_SERVICE_URL", "http://inventory-service:8000")
SALES_SERVICE_URL = os.getenv("SALES_SERVICE_URL", "http://sales-service:8000")
ACCOUNTING_SERVICE_URL = os.getenv("ACCOUNTING_SERVICE_URL", "http://accounting-service:8000")

app = FastAPI(
    title="Reporting & Export Service",
    description="Generate PDF and Excel reports with MinIO storage",
    version="1.0.0"
)


def get_db():
    with get_db_session(SessionFactory) as session:
        yield session


# ============================================
# Schemas
# ============================================

class ReportCreate(BaseModel):
    report_type: models.ReportType
    report_format: models.ReportFormat
    title: str
    description: Optional[str] = None
    report_date: Optional[str] = Field(None, pattern=r'^\d{4}-\d{2}-\d{2}$')
    filters: Optional[Dict[str, Any]] = None
    template_id: Optional[UUID] = None


class ReportResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: UUID
    report_type: models.ReportType
    report_format: models.ReportFormat
    status: models.ReportStatus
    title: str
    description: Optional[str]
    report_date: Optional[str]
    filters: Optional[str]
    file_path: Optional[str]
    file_size: Optional[int]
    file_name: Optional[str]
    user_id: UUID
    generation_time_ms: Optional[int]
    error_message: Optional[str]
    created_at: datetime


class ReportTemplateCreate(BaseModel):
    code: str
    name: str
    description: Optional[str] = None
    report_type: models.ReportType
    report_format: models.ReportFormat
    template_config: Optional[Dict[str, Any]] = None
    default_filters: Optional[Dict[str, Any]] = None
    html_template: Optional[str] = None
    excel_config: Optional[Dict[str, Any]] = None
    is_public: int = 0


class ReportTemplateResponse(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: UUID
    code: str
    name: str
    description: Optional[str]
    report_type: models.ReportType
    report_format: models.ReportFormat
    template_config: Optional[str]
    default_filters: Optional[str]
    is_active: int
    is_public: int
    created_at: datetime


class DashboardData(BaseModel):
    """Dashboard aggregated data"""
    sales_today: int
    sales_this_month: int
    sales_this_year: int
    inventory_value: int
    low_stock_items: int
    pending_sales: int
    tva_this_month: int
    total_customers: int


# ============================================
# Helper Functions
# ============================================

async def fetch_service_data(url: str, headers: Dict[str, str]) -> Optional[Dict]:
    """Fetch data from another service"""
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(url, headers=headers)
            response.raise_for_status()
            return response.json()
    except Exception as e:
        logger.error(f"Error fetching from {url}: {e}")
        return None


def upload_to_minio(file_content: bytes, file_name: str, content_type: str) -> Optional[str]:
    """Upload file to MinIO and return path"""
    if not minio_client:
        logger.warning("MinIO client not available")
        return None

    try:
        file_path = f"{datetime.now().strftime('%Y/%m/%d')}/{file_name}"
        minio_client.put_object(
            MINIO_BUCKET,
            file_path,
            io.BytesIO(file_content),
            length=len(file_content),
            content_type=content_type
        )
        logger.info(f"Uploaded to MinIO: {file_path}")
        return file_path
    except S3Error as e:
        logger.error(f"MinIO upload error: {e}")
        return None


def download_from_minio(file_path: str) -> Optional[bytes]:
    """Download file from MinIO"""
    if not minio_client:
        return None

    try:
        response = minio_client.get_object(MINIO_BUCKET, file_path)
        return response.read()
    except S3Error as e:
        logger.error(f"MinIO download error: {e}")
        return None
    finally:
        if 'response' in locals():
            response.close()
            response.release_conn()


# ============================================
# Excel Generation Functions
# ============================================

def generate_excel_sales_summary(data: Dict, filters: Dict) -> bytes:
    """Generate Excel sales summary report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Summary"

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "Sales Summary Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Filters info
    row = 3
    if filters:
        ws[f'A{row}'] = f"Period: {filters.get('date_from', 'N/A')} to {filters.get('date_to', 'N/A')}"
        row += 2

    # Column headers
    headers = ["Date", "Sale Number", "Customer", "Items", "Amount (FCFA)", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Sample data (in real implementation, fetch from sales service)
    # This is placeholder - actual implementation would fetch real data
    row += 1
    sample_data = data.get('sales', [])
    for sale in sample_data:
        ws.cell(row=row, column=1, value=sale.get('sale_date'))
        ws.cell(row=row, column=2, value=sale.get('sale_number'))
        ws.cell(row=row, column=3, value=sale.get('customer_name'))
        ws.cell(row=row, column=4, value=sale.get('item_count'))
        ws.cell(row=row, column=5, value=sale.get('total_amount', 0) / 100)
        ws.cell(row=row, column=6, value=sale.get('status'))
        row += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.read()


def generate_excel_inventory_status(data: Dict, filters: Dict) -> bytes:
    """Generate Excel inventory status report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Status"

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = "Inventory Status Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Column headers
    row = 3
    headers = ["Product Code", "Product Name", "Type", "Current Stock", "Unit", "Min Level", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    row += 1
    stock_levels = data.get('stock_levels', [])
    for item in stock_levels:
        ws.cell(row=row, column=1, value=item.get('product_code'))
        ws.cell(row=row, column=2, value=item.get('product_name'))
        ws.cell(row=row, column=3, value=item.get('product_type'))
        ws.cell(row=row, column=4, value=float(item.get('current_stock', 0)))
        ws.cell(row=row, column=5, value=item.get('unit'))
        ws.cell(row=row, column=6, value=item.get('min_stock_level'))

        # Status with conditional formatting
        is_below = item.get('is_below_minimum', False)
        status_cell = ws.cell(row=row, column=7)
        status_cell.value = "LOW STOCK" if is_below else "OK"
        if is_below:
            status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            status_cell.font = Font(color="FFFFFF", bold=True)

        row += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.read()


def generate_excel_tva_monthly(data: Dict, filters: Dict) -> bytes:
    """Generate Excel monthly TVA report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "TVA Monthly"

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "Monthly TVA Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Column headers
    row = 3
    headers = ["Month", "TVA Collectée", "TVA Déductible", "TVA Net", "Sales Count", "Purchase Count"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    row += 1
    tva_data = data.get('tva_reports', [])
    for item in tva_data:
        ws.cell(row=row, column=1, value=item.get('fiscal_month'))
        ws.cell(row=row, column=2, value=item.get('tva_collectee', 0) / 100)
        ws.cell(row=row, column=3, value=item.get('tva_deductible', 0) / 100)
        ws.cell(row=row, column=4, value=item.get('tva_net', 0) / 100)
        ws.cell(row=row, column=5, value=item.get('sales_count'))
        ws.cell(row=row, column=6, value=item.get('purchases_count'))
        row += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.read()


# ============================================
# PDF Generation Functions
# ============================================

def generate_pdf_sales_summary(data: Dict, filters: Dict) -> bytes:
    """Generate PDF sales summary report"""
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: A4;
                margin: 2cm;
            }}
            body {{
                font-family: Arial, sans-serif;
                font-size: 10pt;
            }}
            h1 {{
                color: #366092;
                text-align: center;
                border-bottom: 2px solid #366092;
                padding-bottom: 10px;
            }}
            .filters {{
                background-color: #f0f0f0;
                padding: 10px;
                margin: 20px 0;
                border-radius: 5px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 20px;
            }}
            th {{
                background-color: #366092;
                color: white;
                padding: 10px;
                text-align: left;
            }}
            td {{
                padding: 8px;
                border-bottom: 1px solid #ddd;
            }}
            tr:hover {{
                background-color: #f5f5f5;
            }}
            .footer {{
                margin-top: 30px;
                text-align: center;
                font-size: 8pt;
                color: #666;
            }}
        </style>
    </head>
    <body>
        <h1>Sales Summary Report</h1>
        <div class="filters">
            <strong>Period:</strong> {filters.get('date_from', 'N/A')} to {filters.get('date_to', 'N/A')}<br>
            <strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        </div>
        <table>
            <thead>
                <tr>
                    <th>Date</th>
                    <th>Sale Number</th>
                    <th>Customer</th>
                    <th>Items</th>
                    <th>Amount (FCFA)</th>
                    <th>Status</th>
                </tr>
            </thead>
            <tbody>
    """

    sales = data.get('sales', [])
    for sale in sales:
        html_content += f"""
                <tr>
                    <td>{sale.get('sale_date', '')}</td>
                    <td>{sale.get('sale_number', '')}</td>
                    <td>{sale.get('customer_name', '')}</td>
                    <td>{sale.get('item_count', 0)}</td>
                    <td>{sale.get('total_amount', 0) / 100:,.2f}</td>
                    <td>{sale.get('status', '')}</td>
                </tr>
        """

    html_content += """
            </tbody>
        </table>
        <div class="footer">
            Generated by Agricultural Management System - Reporting Service
        </div>
    </body>
    </html>
    """

    pdf_buffer = io.BytesIO()
    HTML(string=html_content).write_pdf(pdf_buffer)
    pdf_buffer.seek(0)
    return pdf_buffer.read()


def generate_pdf_dashboard(data: Dict) -> bytes:
    """Generate PDF dashboard report"""
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: A4;
                margin: 2cm;
            }}
            body {{
                font-family: Arial, sans-serif;
            }}
            h1 {{
                color: #366092;
                text-align: center;
                border-bottom: 3px solid #366092;
                padding-bottom: 15px;
            }}
            .metrics {{
                display: flex;
                flex-wrap: wrap;
                justify-content: space-between;
                margin-top: 30px;
            }}
            .metric-card {{
                width: 45%;
                background-color: #f8f9fa;
                border-left: 4px solid #366092;
                padding: 20px;
                margin-bottom: 20px;
            }}
            .metric-title {{
                color: #666;
                font-size: 12pt;
                margin-bottom: 10px;
            }}
            .metric-value {{
                color: #366092;
                font-size: 24pt;
                font-weight: bold;
            }}
            .footer {{
                margin-top: 50px;
                text-align: center;
                font-size: 9pt;
                color: #666;
            }}
        </style>
    </head>
    <body>
        <h1>Dashboard Report</h1>
        <p style="text-align: center; color: #666;">Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>

        <div class="metrics">
            <div class="metric-card">
                <div class="metric-title">Sales Today</div>
                <div class="metric-value">{data.get('sales_today', 0) / 100:,.2f} FCFA</div>
            </div>
            <div class="metric-card">
                <div class="metric-title">Sales This Month</div>
                <div class="metric-value">{data.get('sales_this_month', 0) / 100:,.2f} FCFA</div>
            </div>
            <div class="metric-card">
                <div class="metric-title">Sales This Year</div>
                <div class="metric-value">{data.get('sales_this_year', 0) / 100:,.2f} FCFA</div>
            </div>
            <div class="metric-card">
                <div class="metric-title">Inventory Value</div>
                <div class="metric-value">{data.get('inventory_value', 0) / 100:,.2f} FCFA</div>
            </div>
            <div class="metric-card">
                <div class="metric-title">Low Stock Items</div>
                <div class="metric-value">{data.get('low_stock_items', 0)}</div>
            </div>
            <div class="metric-card">
                <div class="metric-title">Pending Sales</div>
                <div class="metric-value">{data.get('pending_sales', 0)}</div>
            </div>
            <div class="metric-card">
                <div class="metric-title">TVA This Month</div>
                <div class="metric-value">{data.get('tva_this_month', 0) / 100:,.2f} FCFA</div>
            </div>
            <div class="metric-card">
                <div class="metric-title">Total Customers</div>
                <div class="metric-value">{data.get('total_customers', 0)}</div>
            </div>
        </div>

        <div class="footer">
            Generated by Agricultural Management System - Reporting Service
        </div>
    </body>
    </html>
    """

    pdf_buffer = io.BytesIO()
    HTML(string=html_content).write_pdf(pdf_buffer)
    pdf_buffer.seek(0)
    return pdf_buffer.read()


# ============================================
# Report Endpoints
# ============================================

@app.post("/api/v1/reports", response_model=ReportResponse, status_code=201)
async def create_report(
    report_data: ReportCreate,
    background_tasks: BackgroundTasks,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create and generate a report"""
    start_time = datetime.now()

    # Create report record
    report = models.Report(
        report_type=report_data.report_type,
        report_format=report_data.report_format,
        title=report_data.title,
        description=report_data.description,
        report_date=report_data.report_date,
        filters=json.dumps(report_data.filters) if report_data.filters else None,
        template_id=report_data.template_id,
        user_id=UUID(current_user.user_id),
        status=models.ReportStatus.PROCESSING
    )
    db.add(report)
    db.commit()
    db.refresh(report)

    try:
        # Fetch data based on report type
        filters = report_data.filters or {}
        data = {}

        # Mock data fetching - in production, fetch from actual services
        if report_data.report_type == models.ReportType.SALES_SUMMARY:
            data = {'sales': []}  # Would fetch from sales service
        elif report_data.report_type == models.ReportType.INVENTORY_STATUS:
            data = {'stock_levels': []}  # Would fetch from inventory service
        elif report_data.report_type == models.ReportType.TVA_MONTHLY:
            data = {'tva_reports': []}  # Would fetch from accounting service
        elif report_data.report_type == models.ReportType.DASHBOARD:
            data = {
                'sales_today': 0,
                'sales_this_month': 0,
                'sales_this_year': 0,
                'inventory_value': 0,
                'low_stock_items': 0,
                'pending_sales': 0,
                'tva_this_month': 0,
                'total_customers': 0
            }

        # Generate file based on format
        file_content = None
        content_type = None
        file_extension = None

        if report_data.report_format == models.ReportFormat.PDF:
            if report_data.report_type == models.ReportType.SALES_SUMMARY:
                file_content = generate_pdf_sales_summary(data, filters)
            elif report_data.report_type == models.ReportType.DASHBOARD:
                file_content = generate_pdf_dashboard(data)
            else:
                file_content = generate_pdf_sales_summary(data, filters)  # Default

            content_type = "application/pdf"
            file_extension = "pdf"

        elif report_data.report_format == models.ReportFormat.EXCEL:
            if report_data.report_type == models.ReportType.SALES_SUMMARY:
                file_content = generate_excel_sales_summary(data, filters)
            elif report_data.report_type == models.ReportType.INVENTORY_STATUS:
                file_content = generate_excel_inventory_status(data, filters)
            elif report_data.report_type == models.ReportType.TVA_MONTHLY:
                file_content = generate_excel_tva_monthly(data, filters)
            else:
                file_content = generate_excel_sales_summary(data, filters)  # Default

            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            file_extension = "xlsx"

        if file_content:
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"{report_data.report_type.value}_{timestamp}.{file_extension}"

            # Upload to MinIO
            file_path = upload_to_minio(file_content, file_name, content_type)

            # Update report
            end_time = datetime.now()
            generation_time = int((end_time - start_time).total_seconds() * 1000)

            report.file_path = file_path
            report.file_size = len(file_content)
            report.file_name = file_name
            report.generation_time_ms = generation_time
            report.status = models.ReportStatus.COMPLETED

            # Set expiration (30 days from now)
            expires_at = datetime.now() + timedelta(days=30)
            report.expires_at = expires_at.strftime('%Y-%m-%d %H:%M:%S')

            db.commit()
            db.refresh(report)

            logger.info(f"Report generated successfully", report_id=str(report.id), generation_time=generation_time)

        else:
            report.status = models.ReportStatus.FAILED
            report.error_message = "Failed to generate report content"
            db.commit()

    except Exception as e:
        logger.error(f"Error generating report: {e}", report_id=str(report.id))
        report.status = models.ReportStatus.FAILED
        report.error_message = str(e)
        db.commit()
        db.refresh(report)

    return ReportResponse.model_validate(report)


@app.get("/api/v1/reports", response_model=List[ReportResponse])
async def list_reports(
    report_type: Optional[models.ReportType] = None,
    report_format: Optional[models.ReportFormat] = None,
    status: Optional[models.ReportStatus] = None,
    skip: int = 0,
    limit: int = 100,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """List generated reports"""
    query = db.query(models.Report)

    # Filter by user (non-admin users only see their own reports)
    if not any(role in current_user.roles for role in [Roles.ADMIN, Roles.COMPTABLE]):
        query = query.filter_by(user_id=UUID(current_user.user_id))

    if report_type:
        query = query.filter_by(report_type=report_type)
    if report_format:
        query = query.filter_by(report_format=report_format)
    if status:
        query = query.filter_by(status=status)

    reports = query.order_by(models.Report.created_at.desc()).offset(skip).limit(limit).all()
    return [ReportResponse.model_validate(r) for r in reports]


@app.get("/api/v1/reports/{report_id}", response_model=ReportResponse)
async def get_report(
    report_id: UUID,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get report by ID"""
    report = db.query(models.Report).filter_by(id=report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    # Check access
    if not any(role in current_user.roles for role in [Roles.ADMIN, Roles.COMPTABLE]):
        if str(report.user_id) != current_user.user_id:
            raise HTTPException(status_code=403, detail="Access denied")

    return ReportResponse.model_validate(report)


@app.get("/api/v1/reports/{report_id}/download")
async def download_report(
    report_id: UUID,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Download report file"""
    report = db.query(models.Report).filter_by(id=report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    # Check access
    if not any(role in current_user.roles for role in [Roles.ADMIN, Roles.COMPTABLE]):
        if str(report.user_id) != current_user.user_id:
            raise HTTPException(status_code=403, detail="Access denied")

    if not report.file_path:
        raise HTTPException(status_code=404, detail="Report file not found")

    # Download from MinIO
    file_content = download_from_minio(report.file_path)
    if not file_content:
        raise HTTPException(status_code=404, detail="Failed to download report file")

    # Determine content type
    content_type = "application/pdf" if report.report_format == models.ReportFormat.PDF else \
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        io.BytesIO(file_content),
        media_type=content_type,
        headers={
            "Content-Disposition": f"attachment; filename={report.file_name}"
        }
    )


# ============================================
# Dashboard Data Endpoint
# ============================================

@app.get("/api/v1/dashboard", response_model=DashboardData)
async def get_dashboard_data(
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get aggregated dashboard data"""
    # In production, this would fetch from multiple services
    # For now, return mock data

    dashboard_data = DashboardData(
        sales_today=0,
        sales_this_month=0,
        sales_this_year=0,
        inventory_value=0,
        low_stock_items=0,
        pending_sales=0,
        tva_this_month=0,
        total_customers=0
    )

    # TODO: Implement actual data aggregation from services
    # This would involve making HTTP calls to:
    # - Sales service for sales data
    # - Inventory service for stock data
    # - Accounting service for TVA data

    return dashboard_data


# ============================================
# Report Template Endpoints
# ============================================

@app.post("/api/v1/templates", response_model=ReportTemplateResponse, status_code=201)
async def create_template(
    template_data: ReportTemplateCreate,
    current_user=Depends(require_roles([Roles.ADMIN, Roles.GESTIONNAIRE])),
    db: Session = Depends(get_db)
):
    """Create a report template"""
    # Check if code exists
    if db.query(models.ReportTemplate).filter_by(code=template_data.code).first():
        raise HTTPException(status_code=400, detail="Template code already exists")

    template = models.ReportTemplate(
        **template_data.model_dump(exclude={'template_config', 'default_filters', 'excel_config'}),
        template_config=json.dumps(template_data.template_config) if template_data.template_config else None,
        default_filters=json.dumps(template_data.default_filters) if template_data.default_filters else None,
        excel_config=json.dumps(template_data.excel_config) if template_data.excel_config else None,
        created_by_user_id=UUID(current_user.user_id)
    )
    db.add(template)
    db.commit()
    db.refresh(template)

    logger.info("Report template created", template_id=str(template.id), code=template.code)
    return ReportTemplateResponse.model_validate(template)


@app.get("/api/v1/templates", response_model=List[ReportTemplateResponse])
async def list_templates(
    report_type: Optional[models.ReportType] = None,
    is_active: int = 1,
    skip: int = 0,
    limit: int = 100,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """List report templates"""
    query = db.query(models.ReportTemplate).filter_by(is_active=is_active)

    if report_type:
        query = query.filter_by(report_type=report_type)

    templates = query.offset(skip).limit(limit).all()
    return [ReportTemplateResponse.model_validate(t) for t in templates]


@app.get("/api/v1/templates/{template_id}", response_model=ReportTemplateResponse)
async def get_template(
    template_id: UUID,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get template by ID"""
    template = db.query(models.ReportTemplate).filter_by(id=template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    return ReportTemplateResponse.model_validate(template)


# ============================================
# Startup & Shutdown
# ============================================

@app.on_event("startup")
async def startup_event():
    """Initialize service"""
    logger.info("Reporting service starting up")
    if event_publisher:
        event_publisher.connect()


@app.on_event("shutdown")
async def shutdown_event():
    """Cleanup"""
    logger.info("Reporting service shutting down")
    if event_publisher:
        event_publisher.close()


# ============================================
# Health Check
# ============================================

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    health_status = {
        "status": "healthy",
        "service": "reporting-service",
        "minio": "connected" if minio_client else "unavailable"
    }
    return health_status


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
